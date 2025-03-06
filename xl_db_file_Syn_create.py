import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from pickle import NONE
import mysql.connector
from mysql.connector import Error

tree = NONE  # グローバル変数としてtreeを宣言
inventory = {}  # 在庫を保存する辞書

def sync_all_data():
    global inventory
    connection = None  # ここで connection 変数を初期化
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"

    try:
        if not os.path.exists(filename):
            create_excel_file()

        wb = openpyxl.load_workbook(filename)
        ws_stock = wb["在庫データ"]
        ws_sales = wb["売上データ"]
        ws_purchase = wb["仕入れデータ"]

        # データベースに接続して読み込む
        connection = mysql.connector.connect(
            host='localhost',
            database='sales_inventory',
            user='root',
            password=''
        )
        if connection.is_connected():
            cursor = connection.cursor()
            
            # 在庫データを同期
            cursor.execute("SELECT product, quantity FROM stock")
            db_inventory = {product: quantity for (product, quantity) in cursor}
            while ws_stock.max_row > 1:
                ws_stock.delete_rows(2)
            for product, quantity in db_inventory.items():
                ws_stock.append([product, quantity])
                inventory[product] = quantity

            # 売上データを同期
            cursor.execute("SELECT date, product, quantity, price, total, staff FROM sales")
            db_sales = list(cursor)
            while ws_sales.max_row > 1:
                ws_sales.delete_rows(2)
            for row in db_sales:
                ws_sales.append(row)
            
            # 仕入れデータを同期
            cursor.execute("SELECT date, product, quantity, price, total FROM purchase")
            db_purchase = list(cursor)
            while ws_purchase.max_row > 1:
                ws_purchase.delete_rows(2)
            for row in db_purchase:
                ws_purchase.append(row)

            wb.save(filename)
            wb.close()

        print("すべてのデータを同期しました")  # デバッグ用

    except Error as e:
        print(f"データベースの同期に失敗しました: {e}")  # デバッグ用
    except Exception as e:
        print(f"データの同期に失敗しました: {e}")  # デバッグ用
    finally:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()

# 在庫データを同期する関数を定義
"""def sync_inventory():
    global inventory
    connection = None  # ここで connection 変数を初期化
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    try:
        if not os.path.exists(filename):
            create_excel_file()
            # 新しいファイルが作成された場合、データベースの在庫データをクリア
            connection = mysql.connector.connect(
                host='localhost',
                database='sales_inventory',
                user='root',
                password=''
            )
            if connection.is_connected():
                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE stock")
                connection.commit()
                cursor.close()
                connection.close()
            print("新しいExcelファイルを作成し、データベースの在庫データをクリアしました。")

        wb = openpyxl.load_workbook(filename)
        ws = wb["在庫データ"]

        # データベースに接続して読み込む
        connection = mysql.connector.connect(
            host='localhost',
            database='sales_inventory',
            user='root',
            password=''
        )
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT product, quantity FROM stock")
            db_inventory = {product: quantity for (product, quantity) in cursor}

            # Excelファイルの在庫データをクリア
            while ws.max_row > 1:
                ws.delete_rows(2)

            # Excelファイルにデータベースから読み込んだ在庫データを反映
            for product, quantity in db_inventory.items():
                ws.append([product, quantity])
                inventory[product] = quantity

            wb.save(filename)
            wb.close()

        print("在庫データを同期しました")  # デバッグ用

    except Error as e:
        print(f"データベースの同期に失敗しました: {e}")  # デバッグ用
    except Exception as e:
        print(f"在庫データの同期に失敗しました: {e}")  # デバッグ用
    finally:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()"""

            
# 新しいExcelファイルを作成する関数の中でsync_inventoryを呼び出す
def create_excel_file():
    global wb, ws
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    if os.path.exists(filename):
        # ファイルが既に存在する場合は、それを開く
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        # 売上データシート
        ws_sales = wb.active
        ws_sales.title = "売上データ"
        ws_sales.append(["日付", "商品名", "数量", "単価", "合計", "担当者"])
        
        # 在庫データシート
        ws_inventory = wb.create_sheet("在庫データ")
        ws_inventory.append(["商品名", "在庫数", "発注点"])
        
        # 仕入れデータシート
        ws_purchase = wb.create_sheet("仕入れデータ")
        ws_purchase.append(["日付", "商品名", "数量", "単価", "合計"])
        
        wb.save(filename)
        messagebox.showinfo("成功", "Excelファイルが作成されました。")
        
    # すべてのデータを同期
    sync_all_data()
        
    return filename


def save_sale():
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    # 入力値のバリデーション
    if not all([product_entry.get(), quantity_entry.get(), price_entry.get(), staff_entry.get()]):
        messagebox.showerror("エラー", "すべての項目を入力してください")
        return
    
    try:
        quantity = int(quantity_entry.get())
        price = float(price_entry.get())
        if quantity <= 0 or price <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("エラー", "数量と単価は正の数値で入力してください")
        return
    
    # 入力フィールドから値を取得
    date = datetime.now()
    product = product_entry.get()
    quantity = int(quantity_entry.get())
    price = float(price_entry.get())
    staff = staff_entry.get()
    
    if not os.path.exists(filename):
        create_excel_file()
    
    try: 
        # Excelファイルを開く
        wb = openpyxl.load_workbook(filename)
        ws = wb["売上データ"]
        
        # 新しい行に売上データを追加
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=date.strftime('%Y-%m-%d %H:%M:%S'))
        
        ws.cell(row=next_row, column=2, value=product)
        
        ws.cell(row=next_row, column=3, value=quantity)
        
        ws.cell(row=next_row, column=4, value=price)
        
        ws.cell(row=next_row, column=5, value=quantity * price)
        
        ws.cell(row=next_row, column=6, value=staff)
        
        # 在庫データを更新
        ws_stock = wb["在庫データ"]
        # 行ごとにループ処理   
        for row in range(2, ws_stock.max_row + 1):
            if ws_stock.cell(row=row, column=1).value == product:
                current_stock = ws_stock.cell(row=row, column=2).value
                if current_stock < quantity:
                    messagebox.showerror("エラー", "在庫不足です。売上を記録できません。")
                    wb.close()
                    return
                new_stock = current_stock - quantity
                ws_stock.cell(row=row, column=2, value=new_stock)
                break
        
        # 変更を保存
        wb.save(filename)
        wb.close()  # ファイルを確実に閉じる
            
        # データベースに接続
        connection = mysql.connector.connect(
            host='localhost',
            database='sales_inventory',
            user='root',
            password=''
        )
        if connection.is_connected():
            cursor = connection.cursor()
            # 売上データをデータベースに挿入
            insert_sale_query = """
            INSERT INTO sales (date, product, quantity, price, total, staff)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            sale_data = (date.strftime('%Y-%m-%d %H:%M:%S'), product, quantity, price, quantity * price, staff)
            cursor.execute(insert_sale_query, sale_data)
            connection.commit()
            
            # 在庫データを更新
            update_stock_query = """
            UPDATE stock
            SET quantity = quantity - %s
            WHERE product = %s
            """
            cursor.execute(update_stock_query, (quantity, product))
            connection.commit()
        
        messagebox.showinfo("成功", "売上データが保存され、在庫が更新されました。")

    except PermissionError:
        messagebox.showerror("エラー", "Excelファイルが開いています。閉じてから再度試してください。")
    except Error as e:
        messagebox.showerror("エラー", f"データベースの操作に失敗しました: {e}")
    except Exception as e:
        messagebox.showerror("エラー", f"保存に失敗しました: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    
def view_daily_sales():
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    if not os.path.exists(filename):
        messagebox.showinfo("情報", "売上データファイルが見つかりません。")
        return

    wb = openpyxl.load_workbook(filename)
    ws = wb["売上データ"]
    
    all_sales = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"全ての売上データ: {all_sales}")  # デバッグ用

    sales_window = tk.Toplevel(root)
    sales_window.title("売上一覧")
    
    if not all_sales:
        tk.Label(sales_window, text="売上データはありません。").pack(pady=20)
    else:
        tree = ttk.Treeview(sales_window)
        tree["columns"] = ("日付", "商品名", "数量", "単価", "合計", "担当者")
        
        # 左側の余白を消す
        tree['show'] = 'headings'
        
        for col in tree["columns"]:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="e")  # 各列の幅を100ピクセルに設定
        
        # 日付列は少し広めにしておく
        tree.column("日付", width=150, anchor="e")

        # 数値を右寄せにする
        tree.column("数量", anchor="e")
        tree.column("単価", anchor="e")
        tree.column("合計", anchor="e")
        
        # Treeviewの幅をウィンドウに合わせる
        tree.pack(expand=True, fill="both")

        for row in all_sales:
            tree.insert("", tk.END, values=row)
        
        tree.pack(expand=True, fill="both")

        # 総売上の計算と表示
        total_sales = sum(row[4] for row in all_sales)  # 合計列の合計を計算
        tk.Label(sales_window, text=f"総売上: {total_sales}円").pack(pady=10)

    sales_window.geometry("800x400")
    center_window(sales_window)

        
def manage_inventory():
    global tree
    load_inventory()  # 在庫データを読み込む
    
    inventory_window = tk.Toplevel(root)
    inventory_window.title("在庫管理")

    # 在庫一覧を表示するTreeview
    tree = ttk.Treeview(inventory_window, columns=("商品名", "在庫数"))
    tree['show'] = 'headings'
    
    tree.heading("商品名", text="商品名")
    tree.heading("在庫数", text="在庫数")
    tree.column("商品名", width=150)
    tree.column("在庫数", width=100)
    
    # 保存されている在庫データを表示
    for product, stock in inventory.items():
        tree.insert("", tk.END, values=(product, stock))
    
    tree.pack(pady=10)
    
    # 新しい商品を追加する機能
    tk.Label(inventory_window, text="新しい商品:").pack()
    new_product = tk.Entry(inventory_window)
    new_product.pack()
    tk.Label(inventory_window, text="初期在庫数:").pack()
    initial_stock = tk.Entry(inventory_window)
    initial_stock.pack()
    tk.Button(inventory_window, text="商品追加", command=lambda: add_product(new_product.get(), initial_stock.get())).pack()

    # 既存の商品の在庫を更新する機能
    tk.Label(inventory_window, text="在庫更新:").pack()
    update_product = tk.Entry(inventory_window)
    update_product.pack()
    tk.Label(inventory_window, text="新しい在庫数:").pack()
    new_stock = tk.Entry(inventory_window)
    new_stock.pack()
    tk.Button(inventory_window, text="在庫更新", command=lambda: update_stock(update_product.get(), new_stock.get())).pack()


def save_inventory():
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb["在庫データ"]
        
        # 既存のデータをクリア（ヘッダー行は残す）
        while ws.max_row > 1:
            ws.delete_rows(2)
        
        # 新しいデータを書き込む
        for product, stock in inventory.items():
            ws.append([product, stock])
        
        wb.save(filename)
        wb.close()  # ファイルを確実に閉じる
        
        # データベースに接続
        connection = mysql.connector.connect(
            host='localhost',
            database='sales_inventory',
            user='root',
            password=''
        )
        if connection.is_connected():
            cursor = connection.cursor()
            # テーブルをクリア
            cursor.execute("TRUNCATE TABLE stock")
            # 新しいデータを挿入
            for product, stock in inventory.items():
                insert_stock_query = """
                INSERT INTO stock (product, quantity)
                VALUES (%s, %s)
                """
                cursor.execute(insert_stock_query, (product, stock))
            connection.commit()
        
        print("在庫データを保存しました")  # デバッグ用
        
    except PermissionError:
        messagebox.showerror("エラー", "Excelファイルが開いています。閉じてから再度試してください。")
    except Error as e:
        messagebox.showerror("エラー", f"データベースの操作に失敗しました: {e}")
    except Exception as e:
        messagebox.showerror("エラー", f"在庫データの保存に失敗しました: {e}")
    finally:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()

        
def load_inventory():
    global inventory
    connection = None  # ここで connection 変数を初期化
    filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb["在庫データ"]
        
        # 在庫データを読み込む（2行目以降が在庫データ）
        inventory.clear()  # 既存のデータをクリア
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # 商品名と在庫数が存在する場合
                inventory[row[0]] = row[1]
        print("在庫データを読み込みました")  # デバッグ用
    except Exception as e:
        print(f"在庫データの読み込みに失敗しました: {e}")  # デバッグ用
        
        # データベースに接続して読み込む
        connection = mysql.connector.connect(
            host='localhost',
            database='sales_inventory',
            user='root',
            password=''
        )
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT product, quantity FROM stock")
            for (product, quantity) in cursor:
                inventory[product] = quantity
        print("データベースから在庫データを読み込みました")  # デバッグ用

    except Error as e:
        print(f"データベースの読み込みに失敗しました: {e}")  # デバッグ用
        inventory = {}
    except Exception as e:
        print(f"在庫データの読み込みに失敗しました: {e}")  # デバッグ用
        inventory = {}
    finally:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()


def add_product(product, stock):
    global tree  # この行を追加
    # 新しい商品を追加する関数
    if product and stock:
        try:
            stock = int(stock)
            inventory[product] = stock
            tree.insert("", tk.END, values=(product, stock))
            save_inventory()
            messagebox.showinfo("成功", f"{product}を在庫{stock}で追加しました。")
        except ValueError:
            messagebox.showerror("エラー", "在庫は数字で入力して下さい。")
    else:
        messagebox.showerror("エラー", "商品名と在庫数を入力してください。")

        
def update_stock(product, new_stock):
    # 既存の商品の在庫を更新する関数
    if product in inventory:
        try:
            new_stock = int(new_stock)
            inventory[product] = new_stock
            for item in tree.get_children():
                if tree.item(item)["values"][0] == product:
                    tree.item(item, values=(product, new_stock))
            save_inventory()
            messagebox.showinfo("成功", f"{product}の在庫を{new_stock}に更新しました。")
        except ValueError:
            messagebox.showinfo("エラー", "在庫数は数字で入力してください。")
    else:
        messagebox.showinfo("エラー", "その商品は登録されていません。")


def purchase_entry():
    # 仕入れ入力画面を表示する関数
    # 新しいウィンドウを作成
    purchase_window = tk.Toplevel()
    purchase_window.title("仕入れ入力")
    purchase_window.geometry("300x400")
    
    # 入力フィールドの作成
    ttk.Label(purchase_window, text="商品名:").pack(pady=2)
    purchase_product = ttk.Entry(purchase_window)
    purchase_product.pack(pady=2)
    
    ttk.Label(purchase_window, text="仕入れ数").pack(pady=2)
    purchase_quantity = ttk.Entry(purchase_window)
    purchase_quantity.pack(pady=2)
    
    ttk.Label(purchase_window, text="仕入れ単価").pack(pady=2)
    purchase_price = ttk.Entry(purchase_window)
    purchase_price.pack(pady=2)

    def save_purchase():
        # 仕入れデータを保存する関数
        connection = None  # ここで connection 変数を初期化
        try:
            # 入力の取得と検証
            product = purchase_product.get()
            quantity = int(purchase_quantity.get())
            price = float(purchase_price.get())
            
            # ファイルが存在しない場合に新しいファイルを作成
            filename = f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx"
            if not os.path.exists(filename):
                create_excel_file()
            
            # Excelファイルに保存
            wb = openpyxl.load_workbook(f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx")
            
            # 仕入れデータシートが無ければ作成
            if "仕入れデータ" not in wb.sheetnames:
                wb.create_sheet("仕入れデータ")
                ws = wb["仕入れデータ"]
                ws.append(["日付", "商品名", "仕入れ数", "仕入れ単価", "仕入れ金額"])
            else:
                ws = wb["仕入れデータ"]
                
                # データ追加
                date = datetime.now()
                ws.append([
                    date.strftime('%Y-%m-%d %H:%M:%S'),
                    product,
                    quantity,
                    price,
                    quantity * price
                    ])
                
                # 在庫数の更新
                if product in inventory:
                    inventory[product] += quantity
                else:
                    inventory[product] = quantity
                    
                wb.save(f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx")
                wb.close()
                
                # データベースに接続
                connection = mysql.connector.connect(
                    host='localhost',
                    database='sales_inventory',
                    user='root',
                    password=''
                )
                if connection.is_connected():
                    cursor = connection.cursor()
                    # 仕入れデータをデータベースに挿入
                    insert_purchase_query = """
                    INSERT INTO purchase (date, product, quantity, price, total)
                    VALUES (%s, %s, %s, %s, %s)
                    """
                    purchase_data = (date.strftime('%Y-%m-%d %H:%M:%S'), product, quantity, price, quantity * price)
                    cursor.execute(insert_purchase_query, purchase_data)
                    connection.commit()
                
                # 在庫データの保存
                save_inventory()
                
                messagebox.showinfo("成功", "仕入れデータを保存しました。")
                purchase_window.destroy()
            
        except ValueError:
            messagebox.showerror("エラー", "数値を正しく入力してください。")
        except Error as e:
            messagebox.showerror("エラー", f"データベースの操作に失敗しました: {e}")
        except Exception as e:
            messagebox.showerror("エラー", f"保存に失敗しました: {e}")
            
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
            
        # 保存ボタン
    ttk.Button(purchase_window, text="保存", command=save_purchase).pack(pady=10)


# 集計表生成の関数
def generate_report():
    """売上と仕入れの集計表を生成する関数"""
    try:
        # Excelファイルを読み込み
        wb = openpyxl.load_workbook(f"sales_{datetime.now().strftime('%Y%m%d')}.xlsx")
        
        # 新しいウィンドウを作成
        report_window = tk.Toplevel()
        report_window.title("集計表")
        report_window.geometry("600x400")

        # Treeviewの作成
        tree = ttk.Treeview(report_window, columns=("商品名", "売上数", "売上金額", "仕入数", "仕入金額", "利益"))
        tree.heading("商品名", text="商品名")
        tree.heading("売上数", text="売上数")
        tree.heading("売上金額", text="売上金額")
        tree.heading("仕入数", text="仕入数")
        tree.heading("仕入金額", text="仕入金額")
        tree.heading("利益", text="利益")

        # 列の設定
        tree.column("#0", width=0, stretch=tk.NO)
        for col in ("商品名", "売上数", "売上金額", "仕入数", "仕入金額", "利益"):
            tree.column(col, anchor="e" if col != "商品名" else "w", width=100)

        # データの集計
        sales_data = {}  # 商品ごとの売上データ
        purchase_data = {}  # 商品ごとの仕入れデータ

        # 売上データの集計
        ws_sales = wb["売上データ"]
        for row in ws_sales.iter_rows(min_row=2):
            if not row[0].value:  # 空行をスキップ
                continue
            product = row[1].value
            quantity = row[2].value
            total = row[4].value
            
            if product not in sales_data:
                sales_data[product] = {"quantity": 0, "total": 0}
            sales_data[product]["quantity"] += quantity
            sales_data[product]["total"] += total

        # 仕入れデータの集計
        if "仕入れデータ" in wb.sheetnames:
            ws_purchase = wb["仕入れデータ"]
            for row in ws_purchase.iter_rows(min_row=2):
                if not row[0].value:  # 空行をスキップ
                    continue
                product = row[1].value
                quantity = row[2].value
                total = row[4].value
                
                if product not in purchase_data:
                    purchase_data[product] = {"quantity": 0, "total": 0}
                purchase_data[product]["quantity"] += quantity
                purchase_data[product]["total"] += total

        # 集計データをTreeviewに表示
        for product in set(list(sales_data.keys()) + list(purchase_data.keys())):
            sales_qty = sales_data.get(product, {"quantity": 0, "total": 0})["quantity"]
            sales_total = sales_data.get(product, {"quantity": 0, "total": 0})["total"]
            purchase_qty = purchase_data.get(product, {"quantity": 0, "total": 0})["quantity"]
            purchase_total = purchase_data.get(product, {"quantity": 0, "total": 0})["total"]
            profit = sales_total - purchase_total

            tree.insert("", tk.END, values=(
                product,
                f"{sales_qty:,}",
                f"¥{sales_total:,.0f}",
                f"{purchase_qty:,}",
                f"¥{purchase_total:,.0f}",
                f"¥{profit:,.0f}"
            ))

        # 合計行の追加
        total_sales_qty = sum(data["quantity"] for data in sales_data.values())
        total_sales_amount = sum(data["total"] for data in sales_data.values())
        total_purchase_qty = sum(data["quantity"] for data in purchase_data.values())
        total_purchase_amount = sum(data["total"] for data in purchase_data.values())
        total_profit = total_sales_amount - total_purchase_amount

        tree.insert("", tk.END, values=(
            "合計",
            f"{total_sales_qty:,}",
            f"¥{total_sales_amount:,.0f}",
            f"{total_purchase_qty:,}",
            f"¥{total_purchase_amount:,.0f}",
            f"¥{total_profit:,.0f}"
        ))

        tree.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        wb.close()

    except Exception as e:
        messagebox.showerror("エラー", f"集計表の生成に失敗しました: {e}")


def auto_order():
    # 自動発注伝票生成の関数
    orders = []
    for product, stock in inventory.items():
        if stock < 10:  # 発注点（例：10個）を下回ったら
            orders.append(f"{product}: あと{10-stock}個必要")
    
    if orders:
        order_window = tk.Toplevel(root)
        order_window.title("発注リスト")
        for order in orders:
            ttk.Label(order_window, text=order).pack(pady=5)
    else:
        messagebox.showinfo("情報", "発注が必要な商品はありません")


# メインウィンドウで呼び出す
root = tk.Tk()
root.title("売上・在庫管理システム")
root.geometry("400x650")

style = ttk.Style()
style.configure('TButton', background='#4CAF50', foreground='black', padding=5)
style.configure('TEntry', background='#E0E0E0', padding=5)


# ウィンドウを画面中央に配置
def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'+{x}+{y}')

 
# メインウィンドウの設定後に呼び出す
root.update_idletasks()  # ウィンドウを更新
center_window(root)  # 中央に配置   
    
create_file_button = ttk.Button(root, text="Excelファイル作成", command=create_excel_file, style='TButton')
create_file_button.pack(fill=tk.X, padx=10, pady=5)

ttk.Label(root, text="商品名:").pack(fill=tk.X, padx=10, pady=2)
product_entry = ttk.Entry(root, style='TEntry')
product_entry.pack(fill=tk.X, padx=10, pady=2)

ttk.Label(root, text="数量:").pack(fill=tk.X, padx=10, pady=2)
quantity_entry = ttk.Entry(root, style='TEntry')
quantity_entry.pack(fill=tk.X, padx=10, pady=2)

ttk.Label(root, text="単価:").pack(fill=tk.X, padx=10, pady=2)
price_entry = ttk.Entry(root, style='TEntry')
price_entry.pack(fill=tk.X, padx=10, pady=2)

ttk.Label(root, text="担当者:").pack(fill=tk.X, padx=10, pady=2)
staff_entry = ttk.Entry(root, style='TEntry')
staff_entry.pack(fill=tk.X, padx=10, pady=2)

save_button = ttk.Button(root, text="売上保存", command=save_sale, style='TButton')
save_button.pack(fill=tk.X, padx=10, pady=15)

view_sales_button = ttk.Button(root, text="本日の売上一覧", command=view_daily_sales, style='TButton')
view_sales_button.pack(fill=tk.X, padx=10, pady=15)

inventory_button = ttk.Button(root, text="在庫管理", command=manage_inventory, style='TButton')
inventory_button.pack(fill=tk.X, padx=10, pady=15)

purchase_button = ttk.Button(root, text="仕入れ入力", command=purchase_entry, style='TButton')
purchase_button.pack(fill=tk.X, padx=10, pady=15)
    
report_button = ttk.Button(root, text="集計表生成", command=generate_report, style='TButton')
report_button.pack(fill=tk.X, padx=10, pady=15)

auto_order_button = ttk.Button(root, text="自動発注", command=auto_order, style='TButton')
auto_order_button.pack(fill=tk.X, padx=10, pady=15)

root.mainloop()
    
